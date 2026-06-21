# -*- coding: utf-8 -*-
"""
Gera a planilha de acompanhamento do Checklist de Implementação Service Cloud
da Wilson Sons, validada contra o repositório force-app e a especificação
funcional V2.0, separando o status por Unidade de Negócio.

Não editar manualmente a planilha gerada — ajuste este script e rode novamente.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

UNIDADES = ["Salvador", "Rio Grande", "Centro Logístico", "Rebocadores"]

# Status padronizados
CONCLUIDO = "Concluído"
PARCIAL = "Parcial"
ANDAMENTO = "Em andamento"
NAO_INICIADO = "Não iniciado"
NA = "N/A"
BLOQUEADO = "Bloqueado"

OBS_ORG_ONLY = "Configuração viva na Org WS_SERVICE, não versionada em force-app — validar diretamente na Org."

# Cada item: (epic, frente, atividade, per_unit, status_comum_ou_dict_por_unidade, evidencia, observacao)
# Quando per_unit=True e status é string única, aplica-se igualmente a todas as unidades
# (a menos que vença override em status_overrides).
ITEMS = []

def add(epic, frente, atividade, per_unit, status, evidencia="", observacao="", status_overrides=None):
    ITEMS.append({
        "epic": epic,
        "frente": frente,
        "atividade": atividade,
        "per_unit": per_unit,
        "status": status,
        "evidencia": evidencia,
        "observacao": observacao,
        "overrides": status_overrides or {},
    })

# =====================================================================
# EPIC 1 — FUNDAÇÃO DO SERVICE CLOUD
# =====================================================================
F = "Fundação Service Cloud e Apps"
add(1, F, "Criar um Lightning App para a unidade de negócio", True, NAO_INICIADO,
    "applications/Atendimento_Tecon_Salvador.app-meta.xml",
    "Apenas o App de Salvador existe no repositório.",
    status_overrides={"Salvador": CONCLUIDO})
add(1, F, "Configurar identidade visual, nome, descrição, ícone e navegação do App", True, NAO_INICIADO,
    "applications/Atendimento_Tecon_Salvador.app-meta.xml",
    "Configurado apenas para Salvador (tabs Case, Categorizacao, GestaoSLA, MessagingSession, Account, Contact, Home, Reports, Dashboards + UtilityBar).",
    status_overrides={"Salvador": CONCLUIDO})
add(1, F, "Disponibilizar recursos mínimos por App (Tela Inicial, Contas, Contatos, Casos, Categorização, Relatórios, Painéis)", True, NAO_INICIADO,
    "applications/Atendimento_Tecon_Salvador.app-meta.xml", "", status_overrides={"Salvador": CONCLUIDO})
add(1, F, "Criar Home Page específica por área (filas, casos pendentes, SLA, volumetria, atalhos)", True, NAO_INICIADO,
    "", "Não foi localizada FlexiPage de Home customizada por unidade no repositório.")
add(1, F, "Criar List Views operacionais por fila, status, prioridade, unidade, SLA e backlog", True, NAO_INICIADO, "",
    "Não localizadas listViews específicas no force-app; validar na Org.")
add(1, F, "Configurar Record Types por unidade de negócio no Case", True, CONCLUIDO,
    "objects/Case/recordTypes/AtendimentoTeconSalvador, AtendimentoTeconRioGrande, AtendimentoCentroLogistico, AtendimentoRebocadores",
    "Os 4 Record Types existem.")
add(1, F, "Configurar Record Types por unidade no objeto Categorização", True, CONCLUIDO,
    "objects/Categorizacao__c/recordTypes/* (4 RTs)", "")
add(1, F, "Garantir que a navegação do App reduza ruído operacional e mostre apenas recursos relevantes para cada área", True, NAO_INICIADO, "", "")
add(1, F, "Validar que os Apps organizam experiência mas não substituem controles reais de segurança", False, NAO_INICIADO, "",
    "Depende da validação de OWD/Sharing (ver frente Segurança).")

# =====================================================================
# EPIC 2 — CONTA, CONTATO E GOVERNANÇA CADASTRAL
# =====================================================================
F = "Conta, Contato e Governança Cadastral"
add(2, F, "Configurar Conta para representar Grupo Econômico ou Conta Unidade Operacional", False, NAO_INICIADO, "",
    "Campo AccountType / DominiosEmail__c não localizados em objects/Account no repositório local.")
add(2, F, "Configurar hierarquia de contas via Parent Account", False, NAO_INICIADO, "", "Padrão nativo, não requer metadado customizado — validar uso na Org.")
add(2, F, "Criar/validar campo de domínios de e-mail na Conta (DominiosEmail__c)", False, NAO_INICIADO, "objects/Account/fields", "Campo não encontrado em force-app/main/default/objects/Account/fields.")
add(2, F, "Garantir que domínios sejam preenchidos apenas na conta principal ou no Grupo Econômico", False, NAO_INICIADO, "", "Depende de Validation Rule, não localizada.")
add(2, F, "Bloquear/orientar para que contas subordinadas não tenham domínio preenchido", False, NAO_INICIADO, "", "")
add(2, F, "Validar regra de identificação automática de cliente por domínio de e-mail", False, NAO_INICIADO, "", "Não há Apex de matching por domínio identificado (WS_EmailToCase usa rota por destinatário, não por domínio de Account).")
add(2, F, "Garantir que contato existente seja associado ao atendimento corretamente", False, PARCIAL, "WS_EmailToCaseThreadingService", "Threading por Case existe; vínculo automático Contact↔Account por domínio não confirmado.")
add(2, F, "Permitir criação/associação de contato novo a partir do domínio identificado", False, NAO_INICIADO, "", "")
add(2, F, "Validar cenário de despachante, transportadora, representante ou intermediário", False, PARCIAL, "Campo Representante__c em Case (ver Modelo de Dados - Case)", "Campo existe; fluxo de validação funcional não confirmado.")
add(2, F, "Criar/validar campo Representante no Case (lookup para Conta)", False, CONCLUIDO, "Representante__c", "")
add(2, F, "Criar regras de deduplicação e saneamento para Conta e Contato", False, NAO_INICIADO, "", "")
add(2, F, "Criar relatórios de qualidade cadastral", False, NAO_INICIADO, "reports/", "Pasta reports/ não encontrada no repositório local.")

# =====================================================================
# EPIC 3 (parte 1) — MODELO DE DADOS - CASE
# =====================================================================
F = "Modelo de Dados - Case"
add(3, F, "Criar campos de categorização: Tipo de Caso, Categoria, Assunto e Subassunto", False, CONCLUIDO,
    "Case.TipoCaso__c, Categoria__c, Assunto__c, Subassunto__c", "")
add(3, F, "Criar campo Unidade de Negócio", False, CONCLUIDO, "Case.UnidadeNegocio__c", "")
add(3, F, "Criar campo Etapa do Atendimento", False, CONCLUIDO, "Case.EtapaAtendimento__c",
    "Picklist com 10 valores (Novo, Em Triagem, Em Atendimento, Aguardando Cliente, Aguardando Área Interna, Em Acompanhamento, Preparando Retorno ao Cliente, Concluído, Cancelado, Aguardando atendimento).")
add(3, F, "Criar campos de distribuição: fila destino, origem da distribuição, flag de distribuição solicitada", False, CONCLUIDO,
    "Case.FilaDestinoDeveloperName__c, FilaDestinoId__c, OrigemDistribuicao__c, DistribuicaoSolicitada__c", "")
add(3, F, "Criar campos de encerramento: motivo, comentário e solução do caso", False, CONCLUIDO,
    "Case.MotivoEncerramento__c, ComentarioEncerramento__c, SolucaoCaso__c", "")
add(3, F, "Criar campos de pesquisa: enviar pesquisa, pesquisa enviada, data/hora de envio", False, CONCLUIDO,
    "Case.EnviarPesquisaSatisfacao__c, PesquisaEnviada__c, DataHoraEnvioPesquisa__c", "")
add(3, F, "Criar campos operacionais (BL, Booking, DI/DUIMP, contêiner, fatura, valor contestado, modalidade, base/localidade)", False, CONCLUIDO,
    "Case.BL__c, Booking__c, DI_DUIMP__c, NumeroConteiner__c, NumeroFatura__c, ValorContestado__c, Modalidade__c, BaseLocalidade__c", "")
add(3, F, "Criar campo Áreas de Atendimento Ativas (milestones concorrentes)", False, CONCLUIDO, "Case.AreasAtendimentoAtivas__c", "")
add(3, F, "Criar campos de controle de reclamação Salvador (Reclamação Gerada, Data/Hora Geração)", True, NAO_INICIADO,
    "Case.ReclamacaoGerada__c, DtGeracaoReclamacao__c", "Campos existem no objeto; aplicável apenas a Salvador.",
    status_overrides={"Salvador": PARCIAL, "Rio Grande": NA, "Centro Logístico": NA, "Rebocadores": NA})
add(3, F, "Validar campo Idioma de Atendimento (cenários internacionais em Rebocadores)", True, NAO_INICIADO,
    "Case.IdiomaAtendimento__c", "Campo existe; validação funcional específica de Rebocadores pendente.",
    status_overrides={"Salvador": NA, "Rio Grande": NA, "Centro Logístico": NA, "Rebocadores": PARCIAL})
add(3, F, "Validar CRUD/FLS dos campos por perfil, permission set e unidade de negócio", True, NAO_INICIADO, "force-app/main/default/profiles/",
    "40 profiles no repo, não auditados campo a campo nesta rodada.")

# =====================================================================
# EPIC 3 (parte 2) — MODELO DE DADOS - CATEGORIZAÇÃO
# =====================================================================
F = "Modelo de Dados - Categorização"
add(3, F, "Criar objeto Categorização", False, CONCLUIDO, "objects/Categorizacao__c", "")
add(3, F, "Criar campos: Unidade de Negócio, Tipo de Caso, Categoria, Assunto, Subassunto, Status, Descrição, Motivo de Inativação", False, CONCLUIDO,
    "Categorizacao__c.UnidadeNegocios__c, TipoCaso__c, Categoria__c, Assunto__c, Subassunto__c, Ativo__c, Descricao__c, MotivoInativacao__c", "")
add(3, F, "Criar campos de distribuição: Distribuir para fila?, Por categorização?, Campo/Valor de Distribuição, Fila, DeveloperName/Id da Fila", False, CONCLUIDO,
    "Categorizacao__c.DistribuirParaFila__c, PorCategorizacao__c, CampoDistribuicao__c, ValorDistribuicao__c, NomeFilaDistribuicao__c, FilaDeveloperName__c, FilaId__c", "")
add(3, F, "Criar Chave Natural", False, CONCLUIDO, "Categorizacao__c.ChaveNatural__c", "")
add(3, F, "Criar Hash da Chave Natural como campo único", False, CONCLUIDO, "Categorizacao__c.ChaveNaturalHash__c", "")
add(3, F, "Garantir que registros equivalentes não possam ser duplicados", False, CONCLUIDO, "CategorizacaoService / CategorizacaoTriggerHandler",
    "Validação de hash único centralizada em Apex Service/Trigger, conforme item 6.3.3 da especificação.")
add(3, F, "Garantir que a fila seja referenciada por DeveloperName e não por Id fixo", False, CONCLUIDO, "FilaDeveloperName__c", "")
add(3, F, "Garantir que o objeto suporte carga por API/Data Loader sem quebrar regras de integridade", False, PARCIAL,
    "CategorizacaoTrigger -> CategorizacaoTriggerHandler -> CategorizacaoService", "Execução server-side existe; teste de carga em massa não confirmado nesta rodada.")
add(3, F, "Restringir manutenção da categorização a usuários autorizados", False, NAO_INICIADO, "", "Permission Set específico de manutenção de Categorização não localizado (apenas GestaoSLAConfigurador/AdminTecnico).")
add(3, F, "Ocultar ou tornar somente leitura os campos técnicos para usuários funcionais", False, NAO_INICIADO, "", "Não confirmado via FLS nesta rodada.")

# =====================================================================
# EPIC 3 (parte 3) — CARGA DA ÁRVORE DE ATENDIMENTO
# =====================================================================
F = "Carga da Árvore de Atendimento"
add(3, F, "Carregar árvore de atendimento da unidade", True, NAO_INICIADO, "", "Carga de dados é atividade de dados na Org, não versionada em force-app.")
add(3, F, "Normalizar valores repetidos ou equivalentes entre português e inglês", False, NAO_INICIADO, "", "")
add(3, F, "Separar corretamente casos como Reclamação ou Incidente em linhas distintas quando necessário", False, NAO_INICIADO, "", "")
add(3, F, "Validar Tipo de Caso, Categoria, Assunto e Subassunto sem duplicidades", False, NAO_INICIADO, "", "Garantido estruturalmente pelo Hash único; validação de carga real pendente.")
add(3, F, "Validar quais categorizações possuem distribuição automática", True, NAO_INICIADO, "", "")
add(3, F, "Validar quais categorizações exigem campos dinâmicos no Case", True, NAO_INICIADO, "", "Dynamic Forms não localizadas no repositório.")
add(3, F, "Validar quais categorizações acionam áreas internas/N3", True, NAO_INICIADO, "", "")
add(3, F, "Validar quais categorizações exigem anexo obrigatório, valor financeiro ou dados operacionais", True, NAO_INICIADO, "Case.AnexoObrigatorio__c", "Campo de controle existe; matriz funcional de regras pendente.")

# =====================================================================
# EPIC 3 (parte 4) — LWC DE MANUTENÇÃO DA CATEGORIZAÇÃO
# =====================================================================
F = "LWC de Manutenção da Categorização (categorizacaoManager)"
add(3, F, "Criar LWC/tela de manutenção da árvore de categorização", False, CONCLUIDO,
    "lwc/gestaoSLAWorkspace (aba 'Category') + GestaoSLAController.createCategoria/updateCategoria/deactivateCategoria/reactivateCategoria",
    "CONCLUÍDO em 21/06 e DEPLOYADO NA ORG WILSON_SERVICE: a aba Category do gestaoSLAWorkspace agora cobre CRUD completo de Categorizacao__c (árvore + distribuição para fila), escopado por Unidade de Negócio via GestaoSLA__c e restrito por Record Type do usuário. O LWC legado 'categorizacaoManagerV2' (Aura override de New/Edit/View de Categorizacao__c, recuperado da Org) foi descontinuado: sua lógica de distribuição (campos picklist do Case, filas por unidade, show/hide condicional) foi portada para este modal, e o New/Edit/View de Categorizacao__c voltou ao padrão (actionOverrides Default em todos os formFactor). Deploy do código novo: 0Afbe00000ADszJCAT (448 componentes, 0 erros, 65/65 testes). Exclusão dos 4 componentes legados na Org: 0Afbe00000ADt8zCAD (4/4 deletados, 0 erros) — ver Deltas/delta_categorizacao_legacy_lwc_removal/. Correção lateral: removida referência órfã do flow 'Periodo_Before_Trigger' (4 customErrorMessages obsoletas de uma versão anterior do flow) nos arquivos de tradução PT/EN, que bloqueava qualquer deploy desses arquivos — pré-existente, não relacionado a esta tarefa.")
add(3, F, "Permitir modo criação e modo edição", False, CONCLUIDO,
    "GestaoSLAController.createCategoria / updateCategoria", "Implementado via gestaoSLAWorkspace, não via LWC dedicado da spec.")
add(3, F, "Identificar Record Types disponíveis para o usuário", False, CONCLUIDO,
    "GestaoSLAHelper.getAllowedUnidadesNegocio() (novo) + GestaoSLAService.resolveAllowedUnidades()/ensureUnidadeAllowed() (novos) aplicados em getBootstrap, getGestaoDetail, getCategorias, createCategoria, updateCategoria, deactivateCategoria, getInactiveCategorias e reactivateCategoria",
    "CORRIGIDO em 20/06 e DEPLOYADO NA ORG WILSON_SERVICE em 21/06 (0Afbe00000ADszJCAT, 0 erros, 65/65 testes): gap de segurança identificado por leitura de código (usuário com a Custom Permission, via qualquer Permission Set, via/editava Categorizacao__c das 4 unidades, independente do Record Type de Case do seu perfil). Implementado: getAllowedUnidadesNegocio() deriva as unidades permitidas a partir de Schema.RecordTypeInfo.isAvailable() em Case + mapeamento ParametrosAtendimento__mdt.CaseRecordTypeDeveloperName__c->UnidadeNegocio__c (via AtendimentoConfigService, já usado por CaseCreationSelector). Admin Técnico (canAdminTechnicalSettings) mantém acesso irrestrito. Testes novos em GestaoSLAHelperTest e GestaoSLAServiceTest (bootstrap restrito, admin irrestrito, sem unidade disponível, createCategoria/getCategorias bloqueados para unidade não permitida).")
add(3, F, "Preencher Unidade de Negócio automaticamente quando houver apenas uma opção", False, CONCLUIDO,
    "GestaoSLADTO.BootstrapResponse.unidadeUnica (novo campo) + lwc/gestaoSLAWorkspace.resolveSelectedGestaoId() (já existente, sem alteração necessária)",
    "CORRIGIDO em 20/06, decorrente do item anterior: o seletor de Gestão de SLA no LWC já era oculto para quem não é Admin Técnico (canShowTechnical) e resolveSelectedGestaoId() já auto-selecionava gestoes[0] — porém antes da correção esse 'gestoes[0]' podia ser de qualquer unidade, pois a lista não era filtrada. Com getBootstrap() agora retornando apenas as unidades permitidas, a auto-seleção passa a ser correta e segura sem qualquer alteração de LWC. Novo campo BootstrapResponse.unidadeUnica disponível para uso futuro no front-end (ex.: exibir texto 'Unidade: Salvador' fixo).")
add(3, F, "Listar dinamicamente campos picklist do objeto Case (para distribuição)", False, CONCLUIDO,
    "lwc/gestaoSLAWorkspace (modal Category) reaproveita CategorizacaoController.getCasePicklistFields() / CategorizacaoService.getCasePicklistFields() (describe de Case)",
    "CORRIGIDO em 21/06: achado-chave — existia uma solução completa para isso em um componente Aura/LWC legado de override de New/Edit/View de Categorizacao__c (categorizacaoManagerOverride -> categorizacaoManagerV2, recuperado da Org WILSON_SERVICE), nunca antes auditado porque vivia em aura/ e não em lwc/lwc-padrão. Em vez de recriar, portamos a chamada: gestaoSLAWorkspace agora importa getCasePicklistFields do CategorizacaoController existente (sem duplicar lógica de describe).")
add(3, F, "Listar dinamicamente valores da picklist selecionada (para distribuição)", False, CONCLUIDO,
    "lwc/gestaoSLAWorkspace.categoriaCasePicklistValueOptions (getter, deriva do campo escolhido em casePicklistFieldsData)",
    "CORRIGIDO em 21/06 — mesma fonte de dados do item anterior; os valores da picklist do campo selecionado vêm aninhados na resposta de getCasePicklistFields(), sem chamada adicional ao servidor.")
add(3, F, "Gravar API Name do campo e API Value do valor", False, CONCLUIDO, "GestaoSLADTO.CategoriaRequest.campoDistribuicao/valorDistribuicao -> GestaoSLAService.applyDistribuicaoRequest() -> Categorizacao__c.CampoDistribuicao__c/ValorDistribuicao__c",
    "CORRIGIDO em 21/06: GestaoSLAController.createCategoria/updateCategoria agora persistem esses campos a partir do modal de Categoria do gestaoSLAWorkspace.")
add(3, F, "Gravar labels funcionais para exibição e relatório", False, CONCLUIDO, "Categorizacao__c.LabelCampoDistribuicao__c / LabelValorDistribuicao__c",
    "CORRIGIDO em 21/06: preenchidos automaticamente pelo CategorizacaoTriggerHandler -> CategorizacaoService.beforeSave/validateAndNormalize (já existente, roda em qualquer insert/update, incluindo os feitos pelo GestaoSLAService) — não foi necessário duplicar essa lógica.")
add(3, F, "Listar apenas filas de Case relacionadas à Unidade de Negócio", False, CONCLUIDO,
    "lwc/gestaoSLAWorkspace reaproveita CategorizacaoController.getQueues(unidadeNegocio) / CategorizacaoService.getQueues(), escopado por this.gestao.unidadeNegocio",
    "CORRIGIDO em 21/06, reaproveitando o mesmo método já existente e testado do módulo de Categorização legado.")
add(3, F, "Ocultar campos de fila quando Distribuir para fila? = Não", False, CONCLUIDO,
    "lwc/gestaoSLAWorkspace.html — template if:true={showCategoriaDistribuicao} / if:true={showCategoriaCampoValor}",
    "CORRIGIDO em 21/06: replicado o mesmo padrão de show/hide do componente legado categorizacaoManagerV2 (showDistribuicao/showCampoValor).")
add(3, F, "Exigir fila quando Distribuir para fila? = Sim e Por categorização? = Sim", False, CONCLUIDO, "CategorizacaoService.validateAndNormalize (validação server-side, já existente) + UI required={categoriaFilaRequired}",
    "CORRIGIDO em 21/06: confirmado por leitura de código que a obrigatoriedade já era garantida no trigger (row.addError('Fila é obrigatória.')) independentemente da UI; GestaoSLAService.createCategoria/updateCategoria agora capturam o DmlException resultante e relançam como FunctionalException amigável (antes propagava stacktrace cru). A UI assistida (combobox de fila) foi adicionada ao modal.")
add(3, F, "Exigir campo, valor e fila quando a distribuição depender de picklist do Case", False, CONCLUIDO, "CategorizacaoService.validateAndNormalize (validação server-side, já existente) + UI required={categoriaCampoValorRequired}",
    "CORRIGIDO em 21/06 — mesma mecânica do item anterior; testes novos em GestaoSLAServiceTest cobrem os cenários de erro amigável (distribuir sem fila / distribuir por campo sem valor).")
add(3, F, "Recalcular Chave Natural e Hash ao salvar", False, CONCLUIDO, "CategorizacaoService/Trigger", "Recálculo garantido server-side independentemente da origem da gravação (inclusive via Setup ou carga de dados).")
add(3, F, "Validar responsividade e padrão visual SLDS", False, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 4 — JORNADA DO CASE
# =====================================================================
F = "Jornada do Case"
add(4, F, "Substituir o botão padrão New do Case por fluxo/tela customizada de categorização inicial", False, NAO_INICIADO,
    "objects/Case/Case.object-meta.xml (actionOverrides)",
    "Confirmado: não há actionOverride para 'New' no Case nem Quick Action equivalente. O backend (CaseCreationController) já está pronto (ver item abaixo); falta apenas o componente de UI e o override do botão.")
add(4, F, "Criar LWC/Controller para categorização inicial na criação do Case", False, PARCIAL,
    "classes/CaseCreationController + CaseCreationService/Selector/Helper/DTO/TestDataFactory (+ Test)",
    "Backend completo e testado (getInitialContext, getTreeOptions, resolveCategorizationSelection, getAvailableQueues, buildDefaultValues) — equivalente funcional ao caseNewCategorization da especificação. NÃO existe LWC consumidor nem override do botão New ainda; lwc/caseRecategorization reaproveita esse mesmo Controller para a tela de recategorização, mas a tela de CRIAÇÃO não tem front-end próprio.")
add(4, F, "Permitir seleção de Tipo de Caso, Categoria, Assunto e Subassunto antes da tela padrão", False, PARCIAL,
    "CaseCreationController.getTreeOptions / CaseCreationService.getTreeOptions", "Lógica de árvore dependente (cascata Tipo→Categoria→Assunto→Subassunto) implementada e testada no backend; falta o LWC consumidor.")
add(4, F, "Abrir a tela padrão de criação do Case com valores já preenchidos", False, PARCIAL,
    "CaseCreationService.buildDefaultValues", "Monta o mapa de valores default (RecordTypeId, UnidadeNegocio__c, TipoCaso__c, Categoria__c, Assunto__c, Subassunto__c, Priority, Origin, AccountId/ContactId/ParentId herdados) pronto para abrir a tela padrão; falta o LWC que chame esse método e faça a navegação para o New padrão com os valores.")
add(4, F, "Implementar decisão Assumir", False, PARCIAL, "CaseCreationService.buildDefaultValues (action == 'ASSUMIR')",
    "Implementado e testado no backend (CaseCreationServiceTest.testBuildDefaultsAssumir): define OwnerId=usuário atual, AcaoPosCategorizacao__c='Assumir', Status='Aberto', EtapaAtendimento__c='Em Atendimento'. Falta o LWC que exponha essa decisão ao usuário.")
add(4, F, "Implementar decisão Distribuir para fila", False, PARCIAL, "CaseCreationService.buildDefaultValues (action == 'DISTRIBUIR')",
    "Implementado e testado (testBuildDefaultsDistribuirManualWithAllowedQueue/...WithoutQueueFails): resolve fila parametrizada pela categorização (resolveCategorizationSelection) ou exige seleção manual entre filas da unidade (getAvailableQueues). Falta o LWC.")
add(4, F, "Implementar decisão Encerrar na criação", False, PARCIAL, "CaseCreationService.buildDefaultValues (action == 'ENCERRAR')",
    "Implementado e testado (testBuildDefaultsEncerrarNormalizesUnit): define OwnerId=usuário atual, Status='Fechado', EtapaAtendimento__c='Concluído'. Falta o LWC.")
add(4, F, "Garantir que o Case mantenha cópia dos valores da árvore sem lookup obrigatório para Categorização", False, CONCLUIDO,
    "Case.Categoria__c, Assunto__c, Subassunto__c, TipoCaso__c (cópia) + Case.Categorizacao__c (lookup opcional) + CaseCreationService preenche cópia e lookup quando há match", "")
add(4, F, "Criar ação Recategorizar Caso", False, CONCLUIDO,
    "lwc/caseRecategorization (Quick Action ScreenAction em Case) + classes/CaseRecategorizationController/Service/Selector/Helper/DTO/TestDataFactory (+ Test)",
    "Suíte completa e testada (CaseRecategorizationServiceTest). Item da auditoria anterior estava incorreto ao classificar como 'não comitado' — a suíte de classes já está versionada; apenas reconfirmar se o LWC em si está commitado no branch atual (estava listado como ?? em sessão anterior — validar com 'git status').")
add(4, F, "Reaproveitar racional da categorização na recategorização", False, CONCLUIDO,
    "lwc/caseRecategorization importa CaseCreationController.resolveCategorizationSelection e CaseCreationController.getAvailableQueues diretamente",
    "Reuso de código confirmado nos imports do componente — exatamente como pedido na especificação 5.5.4.")
add(4, F, "Após recategorização, abrir tela padrão de edição para preenchimento dos campos dinâmicos", False, CONCLUIDO,
    "lwc/caseRecategorization.js usa NavigationMixin para 'standard__recordPage' após CloseActionScreenEvent", "")
add(4, F, "Manter botão Edit padrão do Case", False, CONCLUIDO, "", "Nenhuma evidência de override do Edit padrão.")
add(4, F, "Configurar Dynamic Forms por unidade, tipo, categoria, assunto e subassunto", True, NAO_INICIADO, "", "Nenhum .flexipage com Dynamic Forms localizado; apenas 3 layouts estáticos no repo (AreaParticipante, Entitlement, ParametrosAtendimento).")
add(4, F, "Configurar validações server-side para campos obrigatórios críticos", False, PARCIAL, "CategorizacaoService", "Validações na Categorização existem; validações de campos dinâmicos do Case não confirmadas.")
add(4, F, "Configurar Path de Etapa do Atendimento", False, NAO_INICIADO, "Case.EtapaAtendimento__c", "Campo existe; .path-meta.xml não localizado no repositório.")
add(4, F, "Configurar Status macro do Case: Aberto, Pausado, Fechado e Cancelado", False, CONCLUIDO, "Case.Status (nativo) + CaseMilestoneMacroService", "")

# =====================================================================
# EPIC 5 — EMAIL-TO-CASE CUSTOMIZADO
# =====================================================================
F = "Email-to-Case Customizado"
add(5, F, "Criar Apex Email Service", False, CONCLUIDO, "WS_EmailToCaseInboundHandler", "")
add(5, F, "Criar handler de entrada", False, CONCLUIDO, "WS_EmailToCaseInboundHandler", "")
add(5, F, "Criar orquestrador do processamento", False, CONCLUIDO, "WS_EmailToCaseOrchestrator", "")
add(5, F, "Criar services separados (rota, dedup, threading, Case, owner, EmailMessage, anexos, logs, config)", False, CONCLUIDO,
    "WS_EmailToCaseRouteService, DedupService, ThreadingService, CaseService, OwnerService, EmailMessageService, AttachmentService, LogService, ConfigService", "")
add(5, F, "Criar Custom Metadata de rota de e-mail", False, PARCIAL, "ParametrosAtendimento__mdt (RouteKey__c, RouteOwnerType__c, MatchMode__c, ReopenClosedCase__c, etc.)",
    "A especificação propõe um CMDT dedicado WS_Email_Route__mdt; a implementação consolidou os campos de rota dentro de ParametrosAtendimento__mdt (1 registro por unidade). Funcionalmente equivalente, porém divergente do nome/objeto da especificação — registrar como decisão de arquitetura.")
add(5, F, "Parametrizar rota por e-mail de negócio", False, CONCLUIDO, "ParametrosAtendimento__mdt.BusinessEmail__c", "")
add(5, F, "Parametrizar Record Type, origem, prioridade e owner por rota", False, CONCLUIDO,
    "CaseRecordTypeDeveloperName__c, CaseOrigin__c, Prioridade__c, RouteOwnerType__c/RouteOwnerDeveloperName__c/RouteOwnerUsername__c", "")
add(5, F, "Resolver owner por fila via Group.DeveloperName", False, CONCLUIDO, "WS_EmailToCaseOwnerService", "")
add(5, F, "Resolver owner por usuário via Username", False, CONCLUIDO, "WS_EmailToCaseOwnerService", "")
add(5, F, "Implementar matching por envelope, To, Cc ou qualquer destinatário", False, CONCLUIDO, "ParametrosAtendimento__mdt.MatchMode__c + WS_EmailToCaseRouteService", "")
add(5, F, "Implementar fallback para rota default, com regra de apenas uma rota default ativa", False, CONCLUIDO, "ParametrosAtendimento__mdt.IsDefaultEmailRoute__c", "")
add(5, F, "Implementar deduplicação por rota e MessageIdentifier", False, CONCLUIDO, "WS_EmailToCaseDedupService", "")
add(5, F, "Implementar threading por token e cabeçalhos", False, CONCLUIDO, "WS_EmailToCaseThreadingService", "")
add(5, F, "Criar Case novo quando não houver thread", False, CONCLUIDO, "WS_EmailToCaseCaseService", "")
add(5, F, "Anexar e-mail ao Case aberto quando houver thread", False, CONCLUIDO, "WS_EmailToCaseEmailMessageService", "")
add(5, F, "Reabrir Case fechado quando a rota permitir e estiver dentro da janela", False, CONCLUIDO, "ReopenClosedCase__c, ReopenWindowMinutes__c, ReopenStatus__c", "")
add(5, F, "Criar novo Case relacionado ao original quando a rota exigir", False, CONCLUIDO, "RelateNewCaseToOriginal__c, Case.WS_Original_Case__c / ParentId", "")
add(5, F, "Criar EmailMessage inbound", False, CONCLUIDO, "WS_EmailToCaseEmailMessageService", "")
add(5, F, "Publicar anexos como ContentVersion", False, CONCLUIDO, "WS_EmailToCaseAttachmentService, CreateContentFiles__c", "")
add(5, F, "Registrar erros em Log de Integração", False, CONCLUIDO, "WS_EmailToCaseLogService, PersistEmailAudit__c", "")
add(5, F, "Não persistir log de sucesso ou duplicidade", False, CONCLUIDO, "WS_EmailToCaseLogService", "Conforme decisão funcional documentada na especificação 6.4.10.")
add(5, F, "Criar testes automatizados (criação, append, reabertura, duplicidade, anexos, owner, rotas, erros)", False, PARCIAL,
    "WS_EmailToCaseServiceTest, WS_EmailToCaseConfigServiceTest, WS_EmailToCaseTestDataFactory",
    "Apenas 3 classes de teste para ~16 classes de produção do módulo — cobertura abaixo do recomendado; risco para regressão.")

# =====================================================================
# EPIC 6 (parte 1) — CHAT, WHATSAPP E BOT SALESFORCE
# =====================================================================
F = "Chat, WhatsApp e Bot Salesforce"
add(6, F, "Criar Bot único MVP da Wilson Sons", False, NAO_INICIADO, "bots/", "Pasta force-app/main/default/bots vazia — Bot não versionado localmente; validar se existe apenas na Org.")
add(6, F, "Definir nome final do Bot", False, NAO_INICIADO, "", "Depende de decisão de negócio.")
add(6, F, "Configurar diálogo para criar Caso", False, NAO_INICIADO, "", "")
add(6, F, "Configurar diálogo para consultar Caso existente", False, NAO_INICIADO, "", "")
add(6, F, "Configurar diálogo para falar com atendente", False, PARCIAL, "WS_AgentWork_Assumir_Case_Transbordo.flow-meta.xml (não comitado)",
    "Flow de transbordo existe em working tree (decision 'Case elegível?' valida origem Chat/WhatsApp); diálogo do Bot propriamente dito não confirmado.")
add(6, F, "Configurar entrada por WhatsApp", False, PARCIAL, "ParametrosAtendimento__mdt.WhatsAppChannelDeveloperName__c (não comitado), MessagingChannelDeveloperName__c", "")
add(6, F, "Identificar Unidade de Negócio pelo número/canal de WhatsApp", False, PARCIAL, "ParametrosAtendimento__mdt.WhatsAppChannelDeveloperName__c", "Campo permite mapear canal→unidade; lógica de resolução em runtime não auditada nesta rodada.")
add(6, F, "Identificar cliente no WhatsApp por telefone", False, NAO_INICIADO, "WSWillContextResolverService (em alteração)", "Há serviço de resolução de contexto (Will) em alteração no branch atual; confirmar cobertura de telefone→Contact.")
add(6, F, "Configurar entrada por Chat/Enhanced Chat", False, NAO_INICIADO, "", "")
add(6, F, "Criar pré-chat com Unidade de Negócio, Nome e E-mail", False, NAO_INICIADO, "", "")
add(6, F, "Garantir que as variáveis do pré-chat sejam capturadas corretamente no fluxo", False, NAO_INICIADO, "", "")
add(6, F, "Armazenar Unidade de Negócio na sessão/conversa", False, NAO_INICIADO, "MessagingContextService.cls (em alteração)", "")
add(6, F, "Armazenar CaseId e CaseNumber quando houver criação de Caso", False, NAO_INICIADO, "WSWillCaseCreationService.cls (em alteração)", "")
add(6, F, "Configurar transbordo para Omni-Channel", False, PARCIAL, "WS_AgentWork_Assumir_Case_Transbordo.flow-meta.xml (não comitado), Route_from_Will.flow-meta.xml", "")
add(6, F, "Validar menus distintos por unidade de negócio", True, NAO_INICIADO, "", "")
add(6, F, "Validar fallback quando cliente não for identificado", False, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 6 (parte 2) — OMNI-CHANNEL E OMNI SUPERVISOR
# =====================================================================
F = "Omni-Channel e Omni Supervisor"
add(6, F, "Criar Service Channel para Case", False, NAO_INICIADO, "serviceChannels/", OBS_ORG_ONLY)
add(6, F, "Criar Service Channel para Chat/Mensagens", False, NAO_INICIADO, "serviceChannels/",
    "Conforme PROJECT_INDEX.md, Marllon renomeou o ServiceChannel 'Email' para 'Case/Caso' diretamente na Org em 17/06 — " + OBS_ORG_ONLY)
add(6, F, "Associar filas de atendimento ao Omni-Channel", True, NAO_INICIADO, "", OBS_ORG_ONLY)
add(6, F, "Criar Routing Configurations por fila/canal", True, NAO_INICIADO, "queueRoutingConfigs/",
    "QueueRoutingConfig 'Service_Routing_Case_N1' criado por Marllon em 17/06 diretamente na Org (não retrieved). " + OBS_ORG_ONLY,
    status_overrides={"Centro Logístico": PARCIAL})
add(6, F, "Criar Presence Configurations", False, NAO_INICIADO, "", OBS_ORG_ONLY)
add(6, F, "Criar status Disponível Todos / Disponível Casos / Disponível Chat / Ocupado", False, PARCIAL,
    "ServicePresenceStatus 'Caso_Disponivel' e 'Caso_Messaging_Online' criados por Marllon em 17/06 na Org.", OBS_ORG_ONLY)
add(6, F, "Validar capacidade inicial de 4 chats + 4 casos", False, NAO_INICIADO, "", OBS_ORG_ONLY)
add(6, F, "Testar roteamento por unidade de negócio", True, NAO_INICIADO, "", "")
add(6, F, "Testar roteamento por tipo de canal", False, NAO_INICIADO, "", "")
add(6, F, "Configurar Omni Supervisor", False, NAO_INICIADO, "", OBS_ORG_ONLY)
add(6, F, "Criar visão de backlog por fila", True, NAO_INICIADO, "", "")
add(6, F, "Criar visão de agentes online, presença, capacidade e trabalho atribuído", False, NAO_INICIADO, "", "")
add(6, F, "Segmentar visão de supervisor por unidade de negócio", True, NAO_INICIADO, "", "")
add(6, F, "Validar que supervisores não tenham visibilidade transversal indevida", True, NAO_INICIADO, "", "Depende da frente de Segurança/Sharing.")

# =====================================================================
# EPIC 7 — SLA, ENTITLEMENT PROCESS E ÁREA PARTICIPANTE
# =====================================================================
F = "SLA, Entitlement Process e Área Participante"
add(7, F, "Configurar Business Hours por unidade ou processo", True, NAO_INICIADO,
    "GestaoSLA__c.BusinessHoursName__c referencia BusinessHours por DeveloperName",
    "BusinessHours.settings-meta.xml inclui 'Atendimento Salvador'; Rio Grande/Centro Logístico/Rebocadores constam como 'a criar na org' no PROJECT_INDEX.md.",
    status_overrides={"Salvador": CONCLUIDO, "Rio Grande": NAO_INICIADO, "Centro Logístico": NAO_INICIADO, "Rebocadores": NAO_INICIADO})
add(7, F, "Criar Entitlement Process do Case", True, CONCLUIDO,
    "entitlementProcesses/atendimento salvador_v2, atendimento rio grande_v2, atendimento centro logistico_v2, atendimento rebocadores_v2", "Todas as 4 unidades possuem processo v2 ativo/default.")
add(7, F, "Criar Milestone Categorização Inicial", False, CONCLUIDO, "milestoneTypes/Triagem", "")
add(7, F, "Criar Milestone Tratamento Primário", False, CONCLUIDO, "milestoneTypes/Atendimento", "")
add(7, F, "Criar Milestones de Atendimento de Área Interna por área N3", False, CONCLUIDO, "milestoneTypes/Atendimento N3", "")
add(7, F, "Criar Milestone Retorno ao Cliente", False, CONCLUIDO, "milestoneTypes/Retorno N3", "")
add(7, F, "Criar Milestone SLA Total", False, CONCLUIDO, "milestoneTypes/SLA Total", "")
add(7, F, "Configurar regra para pausar SLA em Aguardando Cliente", False, CONCLUIDO, "CaseAreaParticipantePauseService, Case.DataHoraPausaMilestone__c", "")
add(7, F, "Configurar regra para pausar SLA em Em Acompanhamento, se aplicável", False, PARCIAL, "CaseAreaParticipantePauseService", "Implementado para Aguardando Cliente; cobertura de 'Em Acompanhamento' não confirmada nesta rodada.")
add(7, F, "Configurar conclusão de milestones no fechamento do Case", False, CONCLUIDO, "CaseMilestoneMacroService", "")
add(7, F, "Configurar cancelamento funcional dos milestones em Case cancelado", False, CONCLUIDO, "CaseMilestoneMacroService", "")
add(7, F, "Usar campo Áreas de Atendimento Ativas para acionar milestones concorrentes", False, CONCLUIDO, "Case.AreasAtendimentoAtivas__c", "")
add(7, F, "Criar objeto Área Participante", False, CONCLUIDO, "objects/AreaParticipante__c (~30 campos)", "")
add(7, F, "Criar Área Participante automática para Categorização Inicial", False, CONCLUIDO, "AreaParticipanteMilestoneSyncService", "Sincronizada a partir do CaseMilestone padrão (OrigemSLA__c='Standard').")
add(7, F, "Criar Área Participante automática para Tratamento Primário", False, CONCLUIDO, "AreaParticipanteMilestoneSyncService", "")
add(7, F, "Criar Área Participante automática para Retorno ao Cliente", False, CONCLUIDO, "AreaParticipanteMilestoneSyncService", "")
add(7, F, "Criar Área Participante Interna quando uma área N3 for acionada", False, CONCLUIDO, "AreaParticipanteService.addParticipation/addParticipationBulk", "")
add(7, F, "Relacionar Área Participante ao CaseMilestone quando disponível", False, CONCLUIDO, "AreaParticipante__c.CaseMilestoneId__c", "")
add(7, F, "Criar LWC caseAtuacoesPanel (cards laterais com status/prazo/progresso/vencimento)", False, CONCLUIDO,
    "lwc/caseAreasParticipantesPanel", "Implementado com nome caseAreasParticipantesPanel (equivalente funcional ao caseAtuacoesPanel da especificação).")
add(7, F, "Permitir encerramento manual apenas para Área Participante Interna", False, CONCLUIDO, "AreaParticipanteController.closeParticipation", "")
add(7, F, "Exigir Comentário de Retorno e Solução Tomada no encerramento", False, CONCLUIDO, "AreaParticipanteDTO.CloseRequestDTO", "")
add(7, F, "Bloquear fechamento de Case quando houver Área Participante bloqueante aberta", False, CONCLUIDO, "CaseTriggerHandler.beforeUpdate (Pacote 20)", "")
add(7, F, "Testar criação automática de Área Participante ao criar Case direto em etapa avançada", False, PARCIAL, "AreaParticipanteServiceTest", "Cobertura existe; cenário específico de criação direta em etapa avançada não confirmado.")
add(7, F, "Testar sincronização de CaseMilestone com Área Participante", False, CONCLUIDO, "AreaParticipanteMilestoneSyncService + testes (Pacote 19)", "")

# =====================================================================
# EPIC 8 — CASOS RECORRENTES / AGENDAMENTO
# =====================================================================
F = "Casos Recorrentes / Agendamento"
add(8, F, "Criar objeto Agendamento", False, CONCLUIDO, "objects/Agendamento__c (16 campos)", "")
add(8, F, "Criar botão/action em Conta para gerenciar agendamentos", False, CONCLUIDO,
    "lwc/accountCaseScheduleManager (target lightning__RecordAction em Account) + Account.PermitirAgendamentoRecorrente__c",
    "Correção de auditoria: item estava marcado como não localizado; o componente existe e está exposto como Quick Action (ScreenAction) e também como componente de RecordPage em Account.")
add(8, F, "Criar LWC accountCaseScheduleManager", False, CONCLUIDO, "lwc/accountCaseScheduleManager + lwc/accountCaseScheduleIndicator",
    "Correção de auditoria: ambos os componentes existem. accountCaseScheduleIndicator (não previsto na especificação original) parece ser um indicador complementar de status do agendamento na RecordPage da Account.")
add(8, F, "Permitir múltiplos agendamentos por Conta", False, CONCLUIDO, "Agendamento__c.Conta__c (lookup, não master-detail único)", "")
add(8, F, "Permitir seleção de categorização do Case recorrente", False, CONCLUIDO, "Agendamento__c.TipoCaso__c, Categoria__c, Assunto__c, Subassunto__c", "")
add(8, F, "Permitir descrição da atividade", False, CONCLUIDO, "Agendamento__c.DescricaoAtividade__c", "")
add(8, F, "Permitir proprietário usuário ou fila", False, CONCLUIDO, "Agendamento__c.OwnerUser__c, OwnerQueueDeveloperName__c, OwnerType__c", "")
add(8, F, "Permitir frequência diária/semanal/dias específicos/mensal/trimestral/anual", False, PARCIAL, "Agendamento__c.Frequencia__c + CaseScheduleService.isEligibleToRun",
    "Campo único de frequência; a lógica de elegibilidade (isEligibleToRun) existe e testada, mas não confirmamos nesta rodada se todas as granularidades da spec (ex: dias específicos da semana) estão cobertas pelos valores de picklist — validar matriz de valores de Frequencia__c.")
add(8, F, "Permitir horário previsto e data de início", False, CONCLUIDO, "Agendamento__c.HorarioExecucao__c, DataInicio__c", "")
add(8, F, "Não exigir data de fim", False, CONCLUIDO, "Agendamento__c (sem campo DataFim__c obrigatório)", "")
add(8, F, "Criar Apex Scheduler", False, CONCLUIDO, "classes/CaseScheduleScheduler (Schedulable) + classes/CaseScheduleBatch (Database.Batchable)",
    "Correção de auditoria: item estava classificado como risco crítico ('motor de execução não encontrado'). O Scheduler agenda o Batch, que carrega Agendamento__c ativos e delega a CaseScheduleService.processSchedules — suíte completa com CaseScheduleSchedulerTest e CaseScheduleBatchTest.")
add(8, F, "Criar service de cálculo de elegibilidade", False, CONCLUIDO, "classes/CaseScheduleService.isEligibleToRun + CaseScheduleServiceTest", "")
add(8, F, "Não gerar Case em sábado/domingo", False, CONCLUIDO, "CaseScheduleService.isWeekend(today) combinado com Agendamento__c.PularFimDeSemana__c",
    "Confirmado no código (linha ~91 de CaseScheduleService): 'if (schedule.PularFimDeSemana__c == true && isWeekend(today))' pula a execução.")
add(8, F, "Registrar última e próxima execução", False, PARCIAL, "Agendamento__c.UltimoErroExecucao__c",
    "Não há campos explícitos de ÚltimaExecução__c/PróximaExecução__c no objeto; o registro de erro existe (UltimoErroExecucao__c) mas não há rastro de 'última execução com sucesso' nem pré-cálculo de próxima execução — confirmar se isso é coberto de outra forma (ex: log de integração) antes de considerar gap real.")
add(8, F, "Registrar falhas em Log de Integração", False, CONCLUIDO, "classes/CaseScheduleHelper.insertIntegrationLogs / buildErrorLog",
    "Correção de auditoria: estava marcado como dependente do Scheduler 'não localizado'. CaseScheduleHelper monta e insere os logs de erro; chamado por CaseScheduleService.processSchedules.")
add(8, F, "Testar agendamentos por Conta no Centro Logístico", True, PARCIAL,
    "classes/CaseScheduleServiceTest, CaseScheduleBatchTest, CaseScheduleSchedulerTest",
    "Cobertura de teste unitário existe para o motor de agendamento; teste de UAT real em sandbox com conta do Centro Logístico ainda pendente.",
    status_overrides={"Salvador": NA, "Rio Grande": NA, "Rebocadores": NA, "Centro Logístico": PARCIAL})

# =====================================================================
# EPIC 9 — PESQUISA DE SATISFAÇÃO
# =====================================================================
F = "Pesquisa de Satisfação"
add(9, F, "Criar uma Salesforce Survey por Unidade de Negócio", True, NAO_INICIADO,
    "ParametrosAtendimento__mdt.SurveyDeveloperName__c / SurveyDeveloperNameEN__c",
    "Parametrização por unidade existe via CMDT; criação efetiva da Survey no Survey Builder não é versionável em force-app — validar na Org.")
add(9, F, "Criar LWC caseCloseWithSurvey", False, CONCLUIDO, "lwc/caseClosureSurvey + CaseClosureSurveyController", "Implementado com nome caseClosureSurvey.")
add(9, F, "Substituir/acompanhar encerramento por botão customizado", False, CONCLUIDO, "lwc/caseClosureSurvey", "")
add(9, F, "Permitir decisão de envio de pesquisa quando aplicável", False, CONCLUIDO, "Case.EnviarPesquisaSatisfacao__c", "")
add(9, F, "Configurar regra de envio por unidade: opcional ou sempre enviar", True, CONCLUIDO, "ParametrosAtendimento__mdt.SurveyMode__c", "")
add(9, F, "Exibir alerta se o Requisitante já recebeu pesquisa no mesmo dia", False, CONCLUIDO,
    "ParametrosAtendimento__mdt.VerifyDailySurvey__c, ShowDailySurveyIndicator__c, AllowSameDaySurvey__c", "")
add(9, F, "Garantir que o alerta seja informativo e não bloqueante", False, CONCLUIDO, "AllowSameDaySurvey__c", "")
add(9, F, "Atualizar PesquisaEnviada__c", False, CONCLUIDO, "Case.PesquisaEnviada__c", "")
add(9, F, "Atualizar DataHoraEnvioPesquisa__c", False, CONCLUIDO, "Case.DataHoraEnvioPesquisa__c", "")
add(9, F, "Enviar e-mail com template visual Wilson Sons", False, CONCLUIDO,
    "ParametrosAtendimento__mdt.EmailTemplateDeveloperName__c/EN, BrandPrimaryColor__c, BrandSecondaryColor__c, BrandAccentColor__c, BrandLogoResourceName__c", "")
add(9, F, "Validar logo, cores e link da pesquisa", True, NAO_INICIADO, "", "Validação visual manual pendente em UAT.")
add(9, F, "Testar envio para cada uma das quatro áreas", True, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 10 — RECLAMAÇÕES SALVADOR
# =====================================================================
F = "Reclamações Salvador"
add(10, F, "Criar objeto estruturado de Reclamações Salvador", False, CONCLUIDO, "objects/Reclamacao__c (42 campos)",
    "Objeto criado com Record Types para todas as 4 unidades + 'Geral', extrapolando o escopo original (apenas Salvador na especificação).")
add(10, F, "Criar campos compatíveis com a planilha histórica da área", False, CONCLUIDO,
    "Reclamacao__c.Fluig__c, RegistroSE__c, StatusFluig__c, Reincidente__c, VozCliente__c, etc.", "")
add(10, F, "Mapear campos do Case para o objeto Reclamação", False, PARCIAL, "Reclamacao__c.Caso__c (lookup)", "Lookup existe; classe de serviço de automação (ex.: ReclamacaoSalvadorService) não localizada no repositório.")
add(10, F, "Criar automação no fechamento do Case", False, NAO_INICIADO, "Case.ReclamacaoGerada__c, DtGeracaoReclamacao__c", "Campos de controle existem no Case; trigger/flow que materializa o registro de Reclamacao__c não localizado nesta rodada — validar se está implementado via Flow não indexado ou pendente.")
add(10, F, "Condicionar geração a Unidade de Negócio = Salvador", False, NAO_INICIADO, "", "Depende do item anterior (automação de fechamento).")
add(10, F, "Condicionar geração a Tipo de Caso = Reclamação", False, NAO_INICIADO, "", "")
add(10, F, "Criar controle para evitar duplicidade", False, PARCIAL, "Case.ReclamacaoGerada__c (flag de controle)", "")
add(10, F, "Preencher data/hora de geração", False, PARCIAL, "Case.DtGeracaoReclamacao__c", "")
add(10, F, "Criar relatórios por procedência/área responsável/causa/impacto/reincidência/voz do cliente", False, NAO_INICIADO, "reports/", "Pasta de relatórios não encontrada no repositório.")
add(10, F, "Validar que a área não dependa mais de planilha externa para análise gerencial", False, NAO_INICIADO, "", "Depende da automação de geração e dos relatórios.")

# Reclamação é Salvador-specific por desenho — demais unidades N/A nas linhas acima exceto onde já indicado
for it in ITEMS:
    if it["frente"] == F and it["per_unit"] is False:
        it["per_unit"] = True
        it["overrides"].update({"Rio Grande": NA, "Centro Logístico": NA, "Rebocadores": NA})
        # Salvador mantém o status calculado (status "comum")
        it["overrides"]["Salvador"] = it["status"] if "Salvador" not in it["overrides"] else it["overrides"]["Salvador"]

# =====================================================================
# EPIC 11 — SEGURANÇA, COMPARTILHAMENTO E LGPD
# =====================================================================
F = "Segurança, Compartilhamento e LGPD"
add(11, F, "Configurar OWD de Case como privado", False, NAO_INICIADO, "", "Security Settings/OWD não são retrieved por padrão em force-app DX — validar diretamente na Org.")
add(11, F, "Criar papéis por unidade de negócio", True, NAO_INICIADO, "roles/", OBS_ORG_ONLY)
add(11, F, "Garantir hierarquia para supervisores da própria unidade", True, NAO_INICIADO, "", "")
add(11, F, "Criar Permission Sets ou Permission Set Groups por unidade", True, NAO_INICIADO, "permissionsets/",
    "Apenas 4 Permission Sets no repo: GestaoSLAConfigurador, GestaoSLAAdminTecnico, CaseAcompanhamentoOperador, AccountCaseScheduleOperador — nenhum nomeado 'Atendimento <Unidade>' conforme especificação 6.5.4.")
add(11, F, "Criar Permission Set de Supervisor da Unidade", True, NAO_INICIADO, "permissionsets/", "")
add(11, F, "Criar Permission Set de Administração/Governança", False, PARCIAL, "permissionsets/GestaoSLAAdminTecnico.permissionset-meta.xml", "Existe para o módulo de SLA; PS de governança geral da plataforma não confirmado.")
add(11, F, "PROPOSTA (Jean Duarte, 20/06): Sanear o modelo de Permission Sets para 2 PS por Unidade de Negócio (Gestor, Atendente) + 1 PS global de Admin Técnico", False, NAO_INICIADO,
    "permissionsets/ (estado atual: GestaoSLAConfigurador, GestaoSLAAdminTecnico, CaseAcompanhamentoOperador, AccountCaseScheduleOperador — todos globais, nenhum por unidade)",
    "Modelo-alvo proposto pelo cliente: por Unidade de Negócio (Salvador, Rio Grande, Centro Logístico, Rebocadores), apenas 2 Permission Sets ativos — 'Gestor <Unidade>' (categorização, regras de SLA, agendamentos, supervisão) e 'Atendente <Unidade>' (operação do Case/Área Participante do dia a dia) — mais 1 Permission Set único e global 'Admin Técnico' para gestão do ecossistema de atendimento (Custom Metadata de rotas, EntitlementProcess, configurações técnicas, exclusões). Isso simplifica a segregação de acesso e elimina a sobreposição atual de PS técnicos por módulo (GestaoSLAConfigurador/AdminTecnico/CaseAcompanhamentoOperador/AccountCaseScheduleOperador) que hoje não diferencia unidade nem perfil operacional (gestor vs atendente). DESENHO PROPOSTO (pendente de validação e implementação): 1) Mapear cada Custom Permission/objeto/campo/fila hoje espalhado pelos 4 PS técnicos para as 2 personas (Gestor/Atendente) e o Admin Técnico; 2) Consolidar GestaoSLAConfigurador -> dividir entre Gestor <Unidade> (categorização/regras da própria unidade) e Admin Técnico (configuração técnica cross-unidade); 3) CaseAcompanhamentoOperador e AccountCaseScheduleOperador -> incorporar nas permissões de Atendente <Unidade> e Gestor <Unidade> respectivamente; 4) Usar Permission Set Groups por unidade para simplificar atribuição; 5) Validar que a restrição por Record Type implementada em GestaoSLAService (ver itens 'Identificar Record Types disponíveis' / 'Preencher Unidade automaticamente') seja a camada de defesa real — os PS por unidade continuam sendo a camada de concessão de acesso a funcionalidades, não de segregação por dados.")
add(11, F, "Garantir que o App não seja usado como única barreira de segurança", False, NAO_INICIADO, "", "Depende da implementação efetiva de OWD/Sharing/Roles, que não foi confirmada no repositório local.")
add(11, F, "Segregar acesso por papéis, filas, Record Types, objeto e campo", True, NAO_INICIADO, "", "")
add(11, F, "Restringir campos sensíveis do Case", False, NAO_INICIADO, "", "")
add(11, F, "Restringir acesso ao Log de Integração", False, NAO_INICIADO, "LogIntegracao__c", "")
add(11, F, "Restringir acesso ao corpo de e-mail em logs", False, PARCIAL, "WS_EmailToCaseLogService", "Log de erro inclui snapshot do corpo do e-mail por desenho (spec 6.4.10); restrição de FLS sobre esse campo não confirmada.")
add(11, F, "Restringir manutenção de Custom Metadata de rotas", False, NAO_INICIADO, "ParametrosAtendimento__mdt", "")
add(11, F, "Restringir manutenção da Categorização", False, NAO_INICIADO, "", "")
add(11, F, "Ocultar ou deixar somente leitura campos técnicos", False, NAO_INICIADO, "", "")
add(11, F, "Avaliar LGPD para e-mail, telefone, anexos, corpo de e-mail, reclamações e logs", False, NAO_INICIADO, "", "Risco regulatório classificado como Médio/Alto no documento de checklist anexo.")
add(11, F, "Validar CRUD/FLS em Apex", False, PARCIAL, "with sharing em GestaoSLAHealthCheckService e demais services", "Padrão 'with sharing' confirmado em parte das classes; varredura completa não realizada nesta rodada.")
add(11, F, "Testar visibilidade cruzada entre as quatro unidades", True, NAO_INICIADO, "", "")
add(11, F, "Testar acesso de supervisor", True, NAO_INICIADO, "", "")
add(11, F, "Testar acesso administrativo", False, NAO_INICIADO, "", "")
add(11, F, "Registrar exceções formais de compartilhamento", False, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 12 — RELATÓRIOS E INDICADORES
# =====================================================================
F = "Relatórios e Indicadores"
add(12, F, "Criar dashboards por unidade de negócio", True, NAO_INICIADO, "dashboards/", "Pasta dashboards/ não encontrada no repositório local.")
add(12, F, "Criar dashboard consolidado para gestão autorizada", False, NAO_INICIADO, "dashboards/", "")
add(12, F, "Relatório de volume de casos por unidade/origem/tipo/status", True, NAO_INICIADO, "reports/", "Pasta reports/ não encontrada no repositório local.")
add(12, F, "Relatório de casos abertos, pendentes, críticos e fechados", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de backlog por fila/responsável/categoria", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de tempo de resposta e tempo de resolução", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de SLA violado", True, NAO_INICIADO, "reports/", "Campo AreaParticipante__c.ViolouSLA__c existe e pode alimentar o relatório quando criado.")
add(12, F, "Relatório de Área Participante aberta / vencida", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de atendimentos por Conta / Grupo Econômico / Contato", True, NAO_INICIADO, "reports/", "")
add(12, F, "Criar relatórios por área responsável", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de qualidade cadastral", False, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de pesquisa de satisfação", True, NAO_INICIADO, "reports/", "")
add(12, F, "Relatório de reclamações Salvador", True, NAO_INICIADO, "reports/",
    "", status_overrides={"Rio Grande": NA, "Centro Logístico": NA, "Rebocadores": NA})
add(12, F, "Validar Omni Supervisor como ferramenta de acompanhamento em tempo real", False, NAO_INICIADO, "", OBS_ORG_ONLY)

# =====================================================================
# EPIC 13 — DEVOPS, ARQUITETURA E QUALIDADE TÉCNICA
# =====================================================================
F = "DevOps, Arquitetura e Qualidade Técnica"
add(13, F, "Versionar todos os metadados no projeto Salesforce DX", False, CONCLUIDO, "force-app/", "")
add(13, F, "Separar configuração, Apex, LWC, Custom Metadata, objetos, layouts, permissões e automações", False, CONCLUIDO, "force-app/main/default/*", "")
add(13, F, "Aplicar padrão Triscal (Controller/Service/Helper/ServiceAgent/DTO/logs)", False, CONCLUIDO,
    "AreaParticipanteController/Service/Selector/DTO, WS_EmailToCase* (Service/ServiceAgent por camada)", "Padrão consistente nos módulos de AreaParticipante, Email-to-Case, GestaoSLA e Survey.")
add(13, F, "Garantir Apex bulk safe", False, PARCIAL, "AreaParticipanteService.addParticipationBulk", "Bulk safety confirmada nos módulos auditados (AreaParticipante, Categorizacao); varredura total não realizada.")
add(13, F, "Garantir ausência de SOQL em loop", False, PARCIAL, "WSWillCaseCreationSelector (corrige SOQL-in-loop), GestaoSLAHealthCheckService (0 DML, 7 SOQL)", "Correções pontuais documentadas; sem lint completo nesta rodada — recomenda-se rodar Code Analyzer.")
add(13, F, "Garantir ausência de DML em loop", False, PARCIAL, "", "Mesma observação acima.")
add(13, F, "Garantir ausência de lógica em Trigger", False, CONCLUIDO, "CategorizacaoTrigger -> Handler, CaseAfterInsertTrigger -> Handler", "Padrão Trigger->Handler->Service confirmado nos módulos auditados.")
add(13, F, "Garantir ausência de hardcoded ID", False, CONCLUIDO, "Pacote 17A/22 — hardcoded Ids removidos conforme histórico do PROJECT_INDEX.md", "")
add(13, F, "Usar DeveloperName para filas, Record Types e referências migráveis", False, CONCLUIDO,
    "Categorizacao__c.FilaDeveloperName__c, ParametrosAtendimento__mdt.CaseRecordTypeDeveloperName__c/RouteOwnerDeveloperName__c, GestaoSLA__c.EntitlementProcessName__c/BusinessHoursName__c", "")
add(13, F, "Criar testes com Assert.areEqual", False, CONCLUIDO, "*Test.cls em todo o repositório", "")
add(13, F, "Criar testes positivos, negativos e bulk", False, PARCIAL, "AreaParticipanteServiceTest, WS_EmailToCaseServiceTest", "Confirmado nos módulos centrais; cobertura desigual entre módulos (ver item de cobertura abaixo).")
add(13, F, "Validar cobertura mínima e qualidade dos testes", False, BLOQUEADO, "",
    "Levantamento aproximado: ~178 classes de produção vs. ~53 classes de teste no repositório (~30%); módulo Email-to-Case com cobertura especialmente baixa (3 testes para ~16 classes). Requer campanha de testes antes de UAT/cutover.")
add(13, F, "Executar análise estática", False, NAO_INICIADO, "", "Recomenda-se rodar skill running-code-analyzer antes do próximo deploy.")
add(13, F, "Criar plano de deploy", False, PARCIAL, "manifest/retrieve-today-20260618.xml, manifest/retrieve-marllon-20260617.xml", "Manifests de retrieve existem; plano formal de deploy/rollback documentado não localizado.")
add(13, F, "Criar plano de rollback", False, NAO_INICIADO, "", "")
add(13, F, "Separar ativações pós-deploy: Email Service, Omni, Bot, Survey e Schedules", False, NAO_INICIADO, "", "")
add(13, F, "Validar dependências entre metadados", False, NAO_INICIADO, "", "")
add(13, F, "Executar deploy validate antes de produção", False, PARCIAL, "Dry-runs documentados no PROJECT_INDEX.md (ex.: 0Afbe00000AAH8HCAX)", "Prática recorrente confirmada para os últimos pacotes.")
add(13, F, "Executar teste regressivo pós-deploy", False, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 14 (parte 1) — HOMOLOGAÇÃO FUNCIONAL
# =====================================================================
F = "Homologação Funcional"
add(14, F, "Criar roteiro UAT para a unidade", True, NAO_INICIADO, "", "")
add(14, F, "Testar abertura manual de Case", True, NAO_INICIADO, "", "Bloqueado por ausência do LWC caseNewCategorization (ver Jornada do Case).")
add(14, F, "Testar categorização inicial", True, NAO_INICIADO, "", "")
add(14, F, "Testar recategorização", True, ANDAMENTO, "lwc/caseRecategorization (não comitado)", "")
add(14, F, "Testar campos dinâmicos", True, NAO_INICIADO, "", "Dynamic Forms não confirmadas no repositório.")
add(14, F, "Testar distribuição para fila", True, NAO_INICIADO, "", "")
add(14, F, "Testar assumir Case", True, NAO_INICIADO, "", "")
add(14, F, "Testar encerrar na criação", True, NAO_INICIADO, "", "")
add(14, F, "Testar Email-to-Case novo / em Case aberto / em Case fechado", True, PARCIAL, "WS_EmailToCaseServiceTest (cobertura unitária)", "Cobertura unitária existe; testes de UAT ponta a ponta na Org pendentes.")
add(14, F, "Testar anexos de e-mail", True, PARCIAL, "WS_EmailToCaseAttachmentService + testes", "")
add(14, F, "Testar logs de erro", True, PARCIAL, "WS_EmailToCaseLogService + testes", "")
add(14, F, "Testar Bot criando Case / consultando Case / transbordo para atendente", True, NAO_INICIADO, "", "Bloqueado: Bot não localizado no repositório (ver Chat/WhatsApp/Bot).")
add(14, F, "Testar Omni-Channel e capacidade / Omni Supervisor", True, NAO_INICIADO, "", OBS_ORG_ONLY)
add(14, F, "Testar SLA e pausas", True, CONCLUIDO, "CaseAreaParticipantePauseService + AreaParticipanteSLAServiceTest", "Cobertura de testes automatizados confirmada; UAT formal com usuários pendente.")
add(14, F, "Testar Área Participante Interna / fechamento / bloqueio com pendência", True, CONCLUIDO, "AreaParticipanteServiceTest, CaseTriggerHandlerTest (Pacote 20)", "")
add(14, F, "Testar Caso recorrente", True, PARCIAL, "CaseScheduleServiceTest, CaseScheduleBatchTest, CaseScheduleSchedulerTest",
    "Motor de agendamento (CaseScheduleScheduler/Batch/Service/Helper) implementado e com cobertura unitária; falta apenas UAT em sandbox com conta real do Centro Logístico.",
    status_overrides={"Salvador": NA, "Rio Grande": NA, "Rebocadores": NA, "Centro Logístico": PARCIAL})
add(14, F, "Testar pesquisa de satisfação", True, PARCIAL, "CaseClosureSurveyController + testes", "")
add(14, F, "Testar reclamação Salvador", True, NAO_INICIADO, "", "Bloqueado: automação de geração não localizada.",
    status_overrides={"Rio Grande": NA, "Centro Logístico": NA, "Rebocadores": NA})
add(14, F, "Testar segregação de acesso", True, NAO_INICIADO, "", "")
add(14, F, "Coletar evidências de homologação", True, NAO_INICIADO, "", "")

# =====================================================================
# EPIC 14 (parte 2) — CUTOVER, TREINAMENTO E SUSTENTAÇÃO
# =====================================================================
F = "Cutover, Treinamento e Sustentação"
add(14, F, "Criar plano de cutover", False, NAO_INICIADO, "", "")
add(14, F, "Criar checklist pré-go-live", False, NAO_INICIADO, "", "Este documento supre parcialmente essa necessidade.")
add(14, F, "Validar usuários produtivos", True, NAO_INICIADO, "", "")
add(14, F, "Validar filas produtivas", True, NAO_INICIADO, "", "")
add(14, F, "Validar permission sets produtivos", True, NAO_INICIADO, "", "")
add(14, F, "Validar rotas de e-mail produtivas", True, PARCIAL, "ParametrosAtendimento__mdt (1 registro de configuração por unidade)", "")
add(14, F, "Validar canais de Chat/WhatsApp produtivos", True, NAO_INICIADO, "", OBS_ORG_ONLY)
add(14, F, "Validar Bot produtivo", False, NAO_INICIADO, "", "")
add(14, F, "Validar Surveys produtivas", True, NAO_INICIADO, "", "")
add(14, F, "Validar schedules produtivas", True, NAO_INICIADO, "", "Bloqueado: Scheduler de Agendamento não localizado.")
add(14, F, "Treinar usuários da unidade", True, NAO_INICIADO, "", "")
add(14, F, "Treinar supervisores", False, NAO_INICIADO, "", "")
add(14, F, "Entregar runbook operacional", False, NAO_INICIADO, "", "")
add(14, F, "Entregar documentação técnica", False, PARCIAL, "AGENTS.md, AI_HANDLERS.md, docs/PROJECT_INDEX.md", "Documentação técnica de arquitetura/governança existe e é mantida; falta documentação funcional formal de treinamento.")
add(14, F, "Entregar matriz de suporte e responsáveis", False, NAO_INICIADO, "", "")
add(14, F, "Definir período de hypercare", False, NAO_INICIADO, "", "")
add(14, F, "Criar fila de ajustes pós-go-live", False, NAO_INICIADO, "", "")
add(14, F, "Definir processo de sustentação", False, NAO_INICIADO, "", "")

# =====================================================================
# Geração da planilha
# =====================================================================

def build_rows():
    rows = []
    seq = 1
    for it in ITEMS:
        if it["per_unit"]:
            for unidade in UNIDADES:
                status = it["overrides"].get(unidade, it["status"])
                rows.append({
                    "id": f"{seq:04d}",
                    "epic": f"Epic {it['epic']}",
                    "frente": it["frente"],
                    "unidade": unidade,
                    "atividade": it["atividade"],
                    "status": status,
                    "evidencia": it["evidencia"],
                    "observacao": it["observacao"],
                })
                seq += 1
        else:
            rows.append({
                "id": f"{seq:04d}",
                "epic": f"Epic {it['epic']}",
                "frente": it["frente"],
                "unidade": "Comum (todas as unidades)",
                "atividade": it["atividade"],
                "status": it["status"],
                "evidencia": it["evidencia"],
                "observacao": it["observacao"],
            })
            seq += 1
    return rows


STATUS_COLORS = {
    CONCLUIDO: "C6EFCE",
    PARCIAL: "FFEB9C",
    ANDAMENTO: "BDD7EE",
    NAO_INICIADO: "F2F2F2",
    NA: "D9D9D9",
    BLOQUEADO: "F4CCCC",
}
STATUS_FONT_COLORS = {
    CONCLUIDO: "006100",
    PARCIAL: "9C6500",
    ANDAMENTO: "1F4E78",
    NAO_INICIADO: "595959",
    NA: "595959",
    BLOQUEADO: "9C0006",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("ID", 8),
    ("Epic", 10),
    ("Frente", 34),
    ("Unidade de Negócio", 20),
    ("Atividade", 60),
    ("Status", 14),
    ("Ambiente", 12),
    ("Responsável", 16),
    ("Data de Validação", 16),
    ("Evidência (arquivo/observação técnica)", 55),
    ("Observação / Bloqueio", 60),
]


def write_checklist_sheet(wb, rows):
    ws = wb.active
    ws.title = "Checklist"
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for r, row in enumerate(rows, start=2):
        values = [
            row["id"], row["epic"], row["frente"], row["unidade"], row["atividade"],
            row["status"], "", "", "", row["evidencia"], row["observacao"],
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_cell = ws.cell(row=r, column=6)
        fill = STATUS_COLORS.get(row["status"], "FFFFFF")
        font_color = STATUS_FONT_COLORS.get(row["status"], "000000")
        status_cell.fill = PatternFill("solid", fgColor=fill)
        status_cell.font = Font(color=font_color, bold=True)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = len(rows) + 1
    last_col = get_column_letter(len(COLUMNS))
    tab = Table(displayName="ChecklistTable", ref=f"A1:{last_col}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    return ws


def write_resumo_sheet(wb, rows):
    ws = wb.create_sheet("Resumo por Unidade")
    unidades_resumo = UNIDADES + ["Comum (todas as unidades)"]
    statuses = [CONCLUIDO, PARCIAL, ANDAMENTO, NAO_INICIADO, BLOQUEADO, NA]

    ws.cell(row=1, column=1, value="Unidade de Negócio").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    for c, status in enumerate(statuses, start=2):
        cell = ws.cell(row=1, column=c, value=status)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    total_col = len(statuses) + 2
    ws.cell(row=1, column=total_col, value="Total").fill = HEADER_FILL
    ws.cell(row=1, column=total_col).font = HEADER_FONT

    ws.column_dimensions["A"].width = 26
    for c in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    for r, unidade in enumerate(unidades_resumo, start=2):
        ws.cell(row=r, column=1, value=unidade).font = Font(bold=True)
        total = 0
        for c, status in enumerate(statuses, start=2):
            count = sum(1 for row in rows if row["unidade"] == unidade and row["status"] == status)
            cell = ws.cell(row=r, column=c, value=count)
            cell.alignment = Alignment(horizontal="center")
            if count > 0:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
            total += count
        ws.cell(row=r, column=total_col, value=total).font = Font(bold=True)
        ws.cell(row=r, column=total_col).alignment = Alignment(horizontal="center")

    ws.cell(row=len(unidades_resumo) + 3, column=1,
            value="Gerado a partir da validação do checklist anexo (Triscal) contra force-app/main/default e Wilson Sons - Especificação Funcional V2.0.").font = Font(italic=True, size=9)
    ws.cell(row=len(unidades_resumo) + 4, column=1,
            value="Itens 'Comum' representam decisões/arquitetura compartilhadas entre as 4 unidades (não duplicadas por unidade).").font = Font(italic=True, size=9)
    return ws


def write_riscos_sheet(wb):
    ws = wb.create_sheet("Riscos e Gaps Críticos")
    headers = ["Risco", "Impacto", "Evidência no repositório", "Mitigação sugerida"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [32, 36, 50, 50]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28

    riscos = [
("Motor de Agendamento implementado, mas sem agendamento ativo do Scheduler confirmado",
         "CaseScheduleScheduler/Batch/Service estão prontos e testados, mas não há evidência (em force-app) de que o Schedulable foi efetivamente agendado (System.schedule) em produção/sandbox.",
         "classes/CaseScheduleScheduler, CaseScheduleBatch, CaseScheduleService/Helper + testes — agendamento do job (CRON) não é versionável em metadata",
         "Validar diretamente na Org se o job está agendado (Setup > Scheduled Jobs); senão, agendar antes do go-live do Centro Logístico. Correção: auditoria anterior havia classificado esse módulo como 'não localizado' — estava errado, o motor existe e está testado."),
        ("Bot Salesforce não versionado",
         "Criação/consulta de Caso via WhatsApp/Chat e transbordo dependem do Bot.",
         "force-app/main/default/bots vazio",
         "Validar se o Bot existe só na Org (retrieve pendente) ou se ainda não foi construído."),
        ("Front-end de criação manual de Case (tela de categorização inicial) não existe, apesar do backend pronto",
         "Botão New do Case não está substituído pela tela customizada de categorização inicial em nenhuma unidade — usuário ainda cria Case pelo formulário padrão sem a árvore assistida.",
         "classes/CaseCreationController/Service/Selector/Helper/DTO (completo e testado); nenhum LWC consumidor nem actionOverride 'New' em objects/Case/Case.object-meta.xml",
         "Construir o LWC consumindo CaseCreationController.getInitialContext/getTreeOptions/buildDefaultValues e o override do botão New — é o único elo faltante da Jornada do Case. Correção: auditoria anterior classificou o módulo inteiro como 'não iniciado'; na verdade backend está pronto, falta só a tela."),
        ("Cobertura de testes do módulo Email-to-Case baixa (~3 testes / ~16 classes)",
         "Risco de regressão em um módulo crítico de entrada de Casos.",
         "WS_EmailToCaseServiceTest, WS_EmailToCaseConfigServiceTest, WS_EmailToCaseTestDataFactory",
         "Expandir suíte de testes (cenários de rota, dedup, reabertura, erros) antes do próximo deploy."),
        ("Custom Metadata de rota diverge do nome da especificação (WS_Email_Route__mdt)",
         "Divergência entre documentação funcional e implementação real pode confundir times futuros.",
         "Campos de rota consolidados em ParametrosAtendimento__mdt",
         "Atualizar a especificação funcional/PROJECT_INDEX.md para refletir a decisão de arquitetura adotada."),
        ("Apps Lightning, Reports e Dashboards só existem para Salvador",
         "Rio Grande, Centro Logístico e Rebocadores não têm experiência de usuário nem indicadores próprios no repositório.",
         "applications/ contém apenas Atendimento_Tecon_Salvador; reports/ e dashboards/ ausentes",
         "Tratar como gap de paridade entre unidades antes do cutover."),
        ("Automação de geração de Reclamação ao fechar Case (Salvador) não localizada",
         "Objeto Reclamacao__c existe mas pode não ser populado automaticamente, mantendo dependência de planilha externa.",
         "Reclamacao__c.Caso__c existe; ReclamacaoSalvadorService não encontrado",
         "Confirmar se a automação está implementada via Flow não documentado; senão, construir."),
        ("Permission Sets por Unidade de Negócio não encontrados",
         "Segregação de acesso por área pode estar dependendo apenas de Profile/App, contrariando a diretriz de segurança do projeto.",
         "Apenas GestaoSLAConfigurador, GestaoSLAAdminTecnico, CaseAcompanhamentoOperador, AccountCaseScheduleOperador em permissionsets/",
         "Criar Permission Sets/Groups por unidade conforme especificação 6.5.4 antes da homologação de segurança."),
        ("Dynamic Forms / Path de Etapa do Atendimento não confirmados",
         "Campos dinâmicos por categorização e visão de jornada operacional podem não estar implementados.",
         "Nenhum flexipage com Dynamic Forms ou .path-meta.xml localizado",
         "Validar diretamente na Org; se ausente, construir conforme item 5.5.6/5.6.3 da especificação."),
    ]
    for r, (risco, impacto, evidencia, mitigacao) in enumerate(riscos, start=2):
        vals = [risco, impacto, evidencia, mitigacao]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return ws


def main():
    wb = openpyxl.Workbook()
    rows = build_rows()
    write_checklist_sheet(wb, rows)
    write_resumo_sheet(wb, rows)
    write_riscos_sheet(wb)
    out_path = "Wilson_Sons_Checklist_Service_Cloud_Validado.xlsx"
    wb.save(out_path)
    print(f"Gerado: {out_path} ({len(rows)} linhas)")


if __name__ == "__main__":
    main()
